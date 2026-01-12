#!/usr/bin/env python3
"""
Parse Excel files containing asset data and extract asset names and amounts
"""

import json
import sys
from typing import List, Dict, Any
import openpyxl


def parse_excel_file(file_path: str) -> List[Dict[str, Any]]:
    """
    Parse a single Excel file and extract assets

    Args:
        file_path: Path to Excel file

    Returns:
        List of dicts with 'name' and 'amount_krw' keys
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active

    assets = []
    header_row = None
    item_col = None
    name_col = None
    amount_col = None

    # Find header row
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=20), 1):
        for j, cell in enumerate(row):
            if cell.value and "항목" in str(cell.value):
                header_row = i
                item_col = j
            if cell.value and "상품명" in str(cell.value):
                name_col = j
            if cell.value and "금액" in str(cell.value):
                amount_col = j

        if header_row and item_col is not None and name_col is not None and amount_col is not None:
            break

    if not header_row:
        print(f"Warning: Could not find header row in {file_path}", file=sys.stderr)
        return []

    print(f"Found header at row {header_row}: item_col={item_col}, name_col={name_col}, amount_col={amount_col}")

    # Categories to skip (these are category headers, not actual assets)
    CATEGORY_KEYWORDS = [
        "자유입출금",
        "신탁",
        "현금",
        "저축성",
        "전자금융",
        "자산",
        "합계",
        "총",
    ]

    # Parse data rows
    for row in sheet.iter_rows(min_row=header_row + 1):
        item_cell = row[item_col] if item_col < len(row) else None
        name_cell = row[name_col] if name_col < len(row) else None
        amount_cell = row[amount_col] if amount_col < len(row) else None

        # Get values
        item_value = item_cell.value if item_cell else None
        name_value = name_cell.value if name_cell else None
        amount_value = amount_cell.value if amount_cell else None

        # Skip if name is empty
        if not name_value:
            continue

        name = str(name_value).strip()
        if not name:
            continue

        # Skip category headers
        if any(keyword in name for keyword in CATEGORY_KEYWORDS):
            if item_value and str(item_value).strip():
                # This is a category row, skip
                continue

        # Skip rows where amount is 0 or invalid
        try:
            amount = float(amount_value) if amount_value else 0.0
            if amount <= 0:
                print(f"Skipping {name}: amount is {amount}")
                continue
        except (TypeError, ValueError):
            print(f"Skipping {name}: invalid amount {amount_value}")
            continue

        assets.append({
            "name": name,
            "amount_krw": amount
        })

    print(f"Parsed {len(assets)} assets from {file_path}")
    return assets


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 parse_assets.py <excel_file1> [excel_file2] ...")
        sys.exit(1)

    all_assets = []
    for file_path in sys.argv[1:]:
        print(f"\nParsing {file_path}...")
        assets = parse_excel_file(file_path)
        all_assets.extend(assets)

    # Write to JSON
    output_file = "assets_raw.json"
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(all_assets, f, ensure_ascii=False, indent=2)

    print(f"\nTotal assets parsed: {len(all_assets)}")
    print(f"Output written to: {output_file}")

    # Print sample
    print("\nSample assets:")
    for asset in all_assets[:5]:
        print(f"  - {asset['name']}: {asset['amount_krw']:,.0f} KRW")


if __name__ == "__main__":
    main()
